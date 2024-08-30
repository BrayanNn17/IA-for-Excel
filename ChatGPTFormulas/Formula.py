from chatgpt import GetGPTResponse
from tkinter import messagebox
import InterfazMain
import InterfazMasiva
import InterfazUni
import openpyxl
try:
 InterfazMain.AbrirMain()
 print(InterfazMain.opcion_seleccionada)

 if InterfazMain.opcion_seleccionada == 'Formulas unitarias':
   texto=InterfazUni.AbrirInterfaz()

 elif InterfazMain.opcion_seleccionada == 'Formulas Masivas':
   texto=InterfazMasiva.AbrirInterfazMasiva()

 if texto[6] == "U":
   values=GetGPTResponse(texto) 
   print(values)
   if values[0] != 'nan' or values[0] != '' :
    Formula=values[0]
    Formula=Formula.strip()   
    xlsx=values[1][3]
  
    filexslx=openpyxl.load_workbook(xlsx, data_only=True)

    sheets=filexslx.sheetnames
    if values[1][5] == '':
     findSheet=filexslx[sheets[0]]
   
    else:
     findSheet_name=values[1][5]
     findSheet = filexslx[findSheet_name]
    findSheet[values[1][4]]=Formula

    findSheet.sheet_properties.showFormulas = True
  
    filexslx.save(xlsx)
    
   else: 
    raise ValueError("Se cierra ventana de aplicación")
 elif texto[6] == "M":
   
   print(texto)
   xlsx=texto[3]
   
   filexslx=openpyxl.load_workbook(xlsx, data_only=True)

   sheets=filexslx.sheetnames
   if texto[5] == '':
     findSheet=filexslx[sheets[0]]
     
   else:
     findSheet_name=texto[5]
     findSheet = filexslx[findSheet_name]

   start = int(texto[1])
   end = int(texto[2])+1
   for n in range(start,end):
    inputGPT=(texto[0],texto[7]+str(n),texto[8]+str(n),texto[3],texto[4],texto[5],texto[6])
    print(inputGPT)
    values=GetGPTResponse(inputGPT)
    print(values)
    Formula=values[0]
    Formula=Formula.strip()  
    findSheet[texto[4]+str(n)]=Formula
    
   filexslx.save(xlsx)

except Exception as e:
    error_message = str(e)
    print("Error:", error_message)
    if "Worksheet'" and "does not exist." in error_message:
        messagebox.showerror("Error", "Hoja inexistente, por favor vuelva a ejecutar la aplicación")
    elif "'NoneType' object is not subscriptable" in error_message:
        messagebox.showinfo("Información","Se cierra ventana de aplicación")
    else:
        messagebox.showerror("Error", e)


